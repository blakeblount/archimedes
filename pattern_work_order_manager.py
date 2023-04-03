import json
import os
import shutil
import smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from comtypes.client import CreateObject
from docx import Document
from openpyxl import load_workbook
from jira import JIRA

import pattern_work_order_gui as PWOG
import pattern_work_order as PWO
import config

class PatternWorkOrderManager:
    def __init__(self, parent, pattern_work_order_form, config):
        self.config = config

        self.pattern_work_order_form = pattern_work_order_form(parent, self.create_pattern_work_order)

    def get_next_order_number(self):
        # Read the Engineering Log Book Excel file
        workbook = load_workbook(self.log_book_path)
        sheet = workbook.active

        # Find the next available order number
        last_order_number_cell = sheet.cell(row=sheet.max_row, column=1)
        next_order_number = int(last_order_number_cell.value.split('-')[1]) + 1

        return f"P-{next_order_number}"

    def update_log_book(self, pattern_work_order):
        workbook = load_workbook(self.log_book_path)
        sheet = workbook.active
        row = sheet.max_row + 1

        sheet.cell(row=row, column=1, value=pattern_work_order.order_number)
        sheet.cell(row=row, column=2, value=pattern_work_order.part_name)
        sheet.cell(row=row, column=3, value=pattern_work_order.drawing_number)
        sheet.cell(row=row, column=4, value=f"{pattern_work_order.work_requested} {pattern_work_order.pattern_type}")
        sheet.cell(row=row, column=5, value=pattern_work_order.pattern_maker)
        sheet.cell(row=row, column=6, value=pattern_work_order.orderer)
        sheet.cell(row=row, column=7, value=pattern_work_order.issue_date.strftime("%m/%d/%Y"))
        sheet.cell(row=row, column=8, value=pattern_work_order.intended_foundry)

        workbook.save(self.log_book_path)

    def generate_word_file(self, pattern_work_order):
        document = Document(self.template_path)

        # Replace placeholders in the template with pattern_work_order data
        for paragraph in document.paragraphs:
            text = paragraph.text
            text = text.replace("{Order_Number}", pattern_work_order.order_number)
            text = text.replace("{Pattern_Maker}", pattern_work_order.pattern_maker)
            text = text.replace("{Order_Issue_Date}", pattern_work_order.issue_date.strftime("%m/%d/%Y"))
            text = text.replace("{Due_Date}", pattern_work_order.due_date.strftime("%m/%d/%Y"))
            text = text.replace("{Orderer}", pattern_work_order.orderer)
            text = text.replace("{Part_Name}", pattern_work_order.part_name)
            text = text.replace("{Drawing_Number}", pattern_work_order.drawing_number)
            text = text.replace("{Work_Requested}", pattern_work_order.work_requested)
            text = text.replace("{Pattern_Type}", pattern_work_order.pattern_type)
            text = text.replace("{Notes}", pattern_work_order.notes)
            paragraph.text = text

        word_file_name = f"{pattern_work_order.order_number} Pattern Work Order for {pattern_work_order.part_name}.docx"
        word_file_path = os.path.join(self.word_files_path, word_file_name)
        document.save(word_file_path)

        return word_file_path

    def generate_pdf_file(self, pattern_work_order, word_file_path):
        word = CreateObject("Word.Application")
        word.Visible = False

        doc = word.Documents.Open(word_file_path)
        pdf_file_name = f"{pattern_work_order.order_number} Pattern Work Order for {pattern_work_order.part_name}.pdf"
        pdf_file_path = os.path.join(self.pdf_files_path, pdf_file_name)
        doc.SaveAs(pdf_file_path, FileFormat=17)  # 17 represents the PDF file format

        doc.Close()
        word.Quit()

    def send_email(self, pattern_work_order):
        #Determine emails
        orderer_email = self.personal_email_address
        if pattern_work_order.pattern_maker=="Intertool":
           pattern_maker_email = self.intertool_email_address
        elif pattern_work_order.pattern_maker=="Pro Pattern":
           pattern_maker_email = self.pro_pattern_email_address
        
        # Create the email message
        msg = MIMEMultipart()
        msg["From"] = orderer_email
        msg["To"] = pattern_maker_email
        msg["Subject"] = f"{pattern_work_order.order_number} Pattern Work Order for {pattern_work_order.part_name}"

        # Attach the pattern_work_order PDF
        pattern_work_order_pdf_path = os.path.join(self.pdf_files_path, f"{pattern_work_order.order_number} Pattern Work Order for {pattern_work_order.part_name}.pdf")
        self.attach_file_to_email(msg, pattern_work_order_pdf_path)

        # Attach the drawing PDF
        drawing_pdf_path = os.path.join(self.drawings_path, f"{pattern_work_order.drawing_number}.pdf")
        self.attach_file_to_email(msg, drawing_pdf_path)

        # Add the email body
        body = f"Please see attached Pattern Work Order for {pattern_work_order.part_name}. Let me know if you have any questions."
        msg.attach(MIMEText(body, "plain"))

        # Send the email
        with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
            server.starttls()
            server.login(self.personal_email_address, self.personal_email_token)
            server.sendmail(orderer_email, pattern_maker_email, msg.as_string())

    def attach_file_to_email(self, msg, file_path):
        with open(file_path, "rb") as attachment_file:
            attachment = MIMEBase("application", "octet-stream")
            attachment.set_payload(attachment_file.read())
        encoders.encode_base64(attachment)
        attachment.add_header("Content-Disposition", f"attachment; filename={os.path.basename(file_path)}")
        msg.attach(attachment)

    def attach_pattern_work_order_to_jira_issue(self, pattern_work_order):
        # Define Jira instance details
        project_key = "PT"

        # Authenticate with Jira
        auth = (self.jira_username, self.jira_api_token)
        jira = JIRA(self.jira_base_url, basic_auth=auth)

        # Define issue summary format
        issue_summary = f"{pattern_work_order.part_name}+"

        # Find issue with the given summary
        issues = jira.search_issues(f'project="{project_key}" AND summary ~ "{issue_summary}"', maxResults=1)

        if issues:
            issue = issues[0]
            pdf_filename = f"{pattern_work_order.order_number} Pattern Work Order for {pattern_work_order.part_name}"

            if os.path.exists(self.pdf_files_path + "\\" + pdf_filename + ".pdf"):
                # Attach the PDF to the issue
                with open(self.pdf_files_path + "\\" +  pdf_filename + ".pdf", 'rb') as pdf_file:
                    jira.add_attachment(issue, pdf_file, filename=pdf_filename)
                print(f"Successfully attached {pdf_filename} to Jira issue {issue.key}.")
            else:
                print(f"Could not find PDF file '{pdf_filename}'.")
        else:
            print(f"Could not find Jira issue with summary '{issue_summary}' in project '{project_key}'.")

    def create_pattern_work_order(self, user_input):
        order_number = self.get_next_order_number()
        user_input["order_number"] = order_number
        pattern_work_order = PatternWorkOrder.from_user_input(user_input)

        self.update_log_book(pattern_work_order)
        word_file_path = self.generate_word_file(pattern_work_order)
        self.generate_pdf_file(pattern_work_order, word_file_path)
        self.send_email(pattern_work_order)
        #self.attach_pattern_work_order_to_jira_issue(pattern_work_order)

        return pattern_work_order

    def load_config(self):
        with open("config.json", "r", encoding="utf-8") as config_file:
            config = json.load(config_file)

        file_paths = config["file_paths"]
        self.template_path = file_paths["pattern_work_order_template_file"]
        self.word_files_path = file_paths["pattern_work_order_Word_folder"]
        self.pdf_files_path = file_paths["pattern_work_order_PDF_folder"]
        self.drawings_path = file_paths["drawings_folder"]
        self.log_book_path = file_paths["engineering_log_book_file"]

        personal_email_settings = config["personal_email_settings"]
        self.smtp_server = personal_email_settings["smtp_server"]
        self.smtp_port = personal_email_settings["smtp_port"]
        self.personal_email_address = personal_email_settings["personal_email_address"]
        self.personal_email_token = personal_email_settings["personal_email_token"]

        jira_settings = config["jira_settings"]
        self.jira_base_url = jira_settings["jira_base_url"]
        self.jira_username = jira_settings["jira_username"]
        self.jira_api_token = jira_settings["jira_api_token"]

        outgoing_email_addresses = config["outgoing_email_addresses"]
        self.intertool_email_address = outgoing_email_addresses["intertool_email_address"]
        self.pro_pattern_email_address = outgoing_email_addresses["pro_pattern_email_address"]
