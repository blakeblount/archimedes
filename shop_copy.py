# This class is contains the logic of the shop copy itself. 
# There are setter and getter methods for the order number, which is stored here.
# There are three primary methods for creating the shop copy:
# The first queries the Syteline SQL database to get the part numbers, line numbers,
# and quantities for each order. The second organizes them into the correct order
# and combines part numbers that appear on multiple lines. The third looks up the 
# PDF drawings for each part, performs OCR to determine where the Job, Item, and Qty
# fields are on the page, draws the correct values on a new, temporary PDF package, prints the
# package to a physical printer, and deletes the temporary package.

import os
import sys
import csv
import tempfile
from pdf2image import convert_from_path
from PIL import Image, ImageDraw, ImageFont, ImageOps
import img2pdf
import pytesseract
import pyodbc
import openpyxl

pytesseract.pytesseract.tesseract_cmd = os.path.join('tesseract.exe')
#os.environ['TESSDATA_PREFIX'] = os.path.join('Tesseract-OCR', 'tessdata')

# Check if we're running in a PyInstaller bundle
if getattr(sys, 'frozen', False):
    bundle_dir = sys._MEIPASS
else:
    bundle_dir = os.path.dirname(os.path.abspath(__file__))
    bundle_dir = os.path.join(bundle_dir, 'Tesseract-OCR')

# The tessdata directory is in the same directory as this script
tessdata_dir_path = bundle_dir

# Add tessdata_dir_path to the TESSDATA_PREFIX environment variable
os.environ['TESSDATA_PREFIX'] = os.path.join(tessdata_dir_path, 'tessdata')

poppler_path = os.path.join('Poppler', 'poppler-0.68.0', 'bin')

class ShopCopy:
    def __init__(self):
        # Values for the order number (as string) and for the table of parts, line numbers, 
        # and quantities (as list of strings)
        self.order_number = None
        self.order_data_table = None
        self.compression_list = None
        self.comp_code_chart = None
        self.progress_callback = None

    def set_order_number(self, order_number):
        # Sets the order number
        self.order_number = order_number

    def get_order_number(self):
        # Gets the order number
        return self.order_number

    def get_compression_list(self):
        # Gets the list of compression parts
        return self.compression_list

    def get_comp_code_chart(self):
        # Gets the compression code chart
        return self.comp_code_chart
    
    def build_drawing_packages(self, part_num, line_num, qty, conn):
        query = f"""
            SELECT *
            FROM [Drawing Packages]
            WHERE [Item] = '{part_num}';
            """

        cursor = conn.execute(query)
        rows = cursor.fetchall()

        root_drawing = [part_num, line_num, qty]

        if len(rows) == 1:
            return [root_drawing]
        else:
            drawing_list = [root_drawing]
            
            for row in rows:
                if row[1] == root_drawing[0]:
                    continue
                else:
                    sub_part_num = row[1]
                    sub_qty = int(row[2]) * qty
                    sub_drawing = self.build_drawing_packages(sub_part_num, line_num, sub_qty, conn)
                    drawing_list += sub_drawing

            return drawing_list

    def query_customer_order_table(self, server, database, username, password):
        # Open connection to SQL database, query table, put data in a list, close connection, return list

        # Create connection string
        conn_str = (
            r'DRIVER=SQL Server;'
            r'SERVER=' + server + ';'
            r'DATABASE=' + database + ';'
            r'UID=' + username + ';'
            r'PWD=' + password + ';'
        )

        # Define SQL query
        order_num_padded = (' ' * (10 - len(str(self.order_number)))) + str(self.order_number)
#        sql_query = (f"SELECT coi.item AS item, "
#                     f"coi.co_line AS co_line, "
#                     f"coi.qty_ordered AS qty "
#                     f"FROM coitem_mst AS coi "
#                     f"WHERE coi.co_num = '{order_num_padded}' "
#                     f"ORDER BY coi.co_line;")

        sql_query = (f"SELECT coi.item AS item, "
                     f"coi.co_line AS co_line, "
                     f"coi.qty_ordered AS qty "
                     f"FROM coitem_mst AS coi "
                     f"WHERE coi.co_num = '{order_num_padded}' "
                     f"ORDER BY coi.co_line;")

        rows = []
        # Open connection and create cursor
        try:
            with pyodbc.connect(conn_str) as conn:
                with conn.cursor() as cursor:
                    # Execute SQL query
                    cursor.execute(sql_query)

                    # Fetch all rows from query
                    rows = cursor.fetchall()
        except Exception as error:
            print(f"Unable to query SQL database.\nError: {error}")

        # Convert rows to list and return
        query_results = [list(row) for row in rows]

        #drawing_conn_str = (r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};"r"DBQ=C:\\Users\\bblount\\Documents\\Drawing Packages.accdb;") 
        drawing_conn_str = (r"Driver={Microsoft Access Driver (*.mdb, *.accdb)};"r"DBQ=\\sefcordata\shared\Engineering\ENG Projects\Drawing Packages.accdb;") 

        with pyodbc.connect(drawing_conn_str) as drawing_conn:
            drawing_package = []
            for row in query_results:
                part_number, line_item, quantity = row
                drawings = self.build_drawing_packages(part_number, line_item, quantity, drawing_conn)
                drawing_package += drawings

        return drawing_package
        
    def organize_shop_copy_data(self, query_results):
        # This method acquires the data in list form, and reorganizes it so that duplicate parts are combined.

        # Create a dictionary of part numbers where each part number has an associated line item and quantity field.
        part_number_dict = {}

        # For each part number, check to see if it's in the list already. If it is, append its line and quantity
        # data to the respective dictionary entries. If it isn't, create a new dictionary entry for the part number.
        for row in query_results:
            part_number, line_item, quantity = row
            if part_number == ('EXPEDITE FEE' or 'FREIGHT CHARGE'):
                continue
            if part_number in part_number_dict:
                part_number_dict[part_number]['line_items'].append(str(line_item))
                part_number_dict[part_number]['quantities'].append(str(int(quantity)))
            else:
                part_number_dict[part_number] = {
                    'line_items': [str(line_item)],
                    'quantities': [str(int(quantity))]
                }

        # Convert to a list for ease of displaying in the UI and using in the print_shop_copy method
        shop_copy_output_table = []
        for part_number, data in part_number_dict.items():
            line_items = ",".join(data['line_items'])
            quantities = ",".join(data['quantities'])
            shop_copy_output_table.append([part_number, line_items, quantities])
            self.order_data_table = shop_copy_output_table

        return shop_copy_output_table

    def extract_conductor_info_from_chart(self):
        try:
            wb = openpyxl.load_workbook(r"\\sefcordata\shared\Engineering\Conductor Chart.xlsx")
            sheet = wb["Code Chart"]
        except Exception as error:
            print(f"Could not open Conductor Chart. {error}")

        hex_column_index = None
        circ_column_index = None
        size_column_index = None
        strand_column_index = None
        type_column_index = None

        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                if cell.value == "COMP HEX CODE":
                    hex_column_index = cell.column
                elif cell.value == "COMP CD CODE":
                    circ_column_index = cell.column
                elif cell.value == "SIZE":
                    size_column_index = cell.column
                elif cell.value == "STRAND":
                    strand_column_index = cell.column
                elif cell.value == "TYPE":
                    type_column_index = cell.column

        hex_values = []
        circ_values = []
        hex_info_list = []
        circ_info_list = []
        
        # Hex die values
        for row in sheet.iter_rows(min_row=2):
            hex_value = ""
            cell_value = sheet.cell(row=row[0].row, column=hex_column_index).value
            if cell_value:
                hex_value += str(cell_value)
            else:
                hex_value = ""
                continue 
            hex_values.append(hex_value)
            size = str(sheet.cell(row=row[0].row, column=size_column_index).value)
            strand = str(sheet.cell(row=row[0].row, column=strand_column_index).value)
            type_ = str(sheet.cell(row=row[0].row, column=type_column_index).value)
            if type_ == "ACSR" or type_ == "ACAR":
                hex_info_list.append(f"{size} {strand} {type_}")
            else:
                hex_info_list.append(f"{size} {type_}")

        # Circular die values
        for row in sheet.iter_rows(min_row=2):
            circ_value = ""
            cell_value = sheet.cell(row=row[0].row, column=circ_column_index).value
            if cell_value:
                circ_value += str(cell_value)
            else:
                circ_value = ""
                continue 
            circ_values.append(circ_value)
            size = str(sheet.cell(row=row[0].row, column=size_column_index).value)
            strand = str(sheet.cell(row=row[0].row, column=strand_column_index).value)
            type_ = str(sheet.cell(row=row[0].row, column=type_column_index).value)
            if type_ == "ACSR" or type_ == "ACAR":
                circ_info_list.append(f"{size} {strand} {type_}")
            else:
                circ_info_list.append(f"{size} {type_}")

        code_chart_dict = {}

        for code, cable in zip(hex_values, hex_info_list):
            if code in code_chart_dict:
                if cable in code_chart_dict[code]:
                    continue
                else:
                    code_chart_dict[code].append(cable)
            else:
                code_chart_dict[code] = [cable]

        for code, cable in zip(circ_values, circ_info_list):
            if code in code_chart_dict:
                if cable in code_chart_dict[code]:
                    continue
                else:
                    code_chart_dict[code].append(cable)
            else:
                code_chart_dict[code] = [cable]

        return code_chart_dict

    def make_compression_list(self):
        compression_list = {}

        compression_prefixes = ['AL', 'AL2', 'AL3', 'ATCF', 'ATCF2', 'ATCC', 'AS', 'ARS', 'ASPCC', 'QCTHV', 'QCT', 'CL', 'CL2', 'CTCF', 'CTCC', 'CS']

        double_code_prefixes = ['ATCC', 'AS', 'CTCC', 'CS']

        for row in self.order_data_table:
            part_number = row[0]
            prefix, compression_code = part_number.split('-')[0:2]
            if prefix in compression_prefixes:
                if prefix in double_code_prefixes:
                   tap_code = part_number.split('-')[2:3][0]
                   compression_list[part_number] = [compression_code.lstrip('0'), tap_code.lstrip('0')]
                else: 
                   compression_list[part_number] = compression_code.lstrip('0')

        if compression_list:
            self.compression_list = compression_list
        else:
            self.compression_list = None

        if not compression_list:
            return None
        else:
            if self.comp_code_chart == None:
                self.comp_code_chart = self.extract_conductor_info_from_chart()
        return True

    def print_shop_copy(self, drawings_path, compression_list):

        # Create list of marked drawing images to be converted into a PDF shop copy packet
        modified_image_paths = []

        # Iterate through each drawing to be printed
        for i, row in enumerate(self.order_data_table):
            drawing_filename = f"{row[0].rstrip(' ')}.pdf"

            # Replace '/' with '[' in drawing file names per SEFCOR practice
            if "/" in drawing_filename:
                drawing_filename = drawing_filename.replace("/", '[')

            job_text = self.order_number
            item_text = row[1]
            qty_text = row[2]

            quantities = qty_text.split(",")
            if "0" in quantities:
                continue

            img = []

            if os.path.exists(drawings_path + drawing_filename):
                img = convert_from_path(drawings_path + drawing_filename, poppler_path=poppler_path)
                print(f"Drawing {i + 1} of {len(self.order_data_table)} ({drawing_filename}) found")
            else:
                print(f"Drawing {i + 1} of {len(self.order_data_table)} ({drawing_filename}) not found")
                continue

            progress = round((i + 1) / len(self.order_data_table) * 100)

            img = img[0]

            try:

                # This line is what runs OCR on the part drawing.
                ocr_data = pytesseract.image_to_data(img, output_type='dict')

                # These are the strings we are looking for on the drawings.
                job_str = "Job"
                item_str = "Item(s)"
                qty_str = "Qty."

                job_left, job_top, job_width, job_height = None, None, None, None
                item_left, item_top, item_width, item_height = None, None, None, None
                qty_left, qty_top, qty_width, qty_height = None, None, None, None

                # The OCR returns a dictionary with headings for the word found, and then location, height, and width in
                # pixels.
                # Establish Job block position
                for i, text in enumerate(ocr_data['text']):
                    if job_str in text:
                        job_left = ocr_data['left'][i]
                        job_top = ocr_data['top'][i]
                        job_width = ocr_data['width'][i]
                        job_height = ocr_data['height'][i]
                        # print(f"job l: {job_left}")
                        # print(f"job t: {job_top}")
                        # print(f"job w: {job_width}")
                        # print(f"job h: {job_height}")

                # Establish Item block position
                for i, text in enumerate(ocr_data['text']):
                    if item_str in text:
                        item_left = ocr_data['left'][i]
                        item_top = ocr_data['top'][i]
                        item_width = ocr_data['width'][i]
                        item_height = ocr_data['height'][i]
                        # print(f"itm l: {item_left}")
                        # print(f"itm t: {item_top}")
                        # print(f"itm w: {item_width}")
                        # print(f"itm h: {item_height}")

                # Establish Quantity block position
                for i, text in enumerate(ocr_data['text']):
                    if qty_str in text:
                        qty_left = ocr_data['left'][i]
                        qty_top = ocr_data['top'][i]
                        qty_width = ocr_data['width'][i]
                        qty_height = ocr_data['height'][i]
                        # print(f"qty l: {qty_left}")
                        # print(f"qty t: {qty_top}")
                        # print(f"qty w: {qty_width}")
                        # print(f"qty h: {qty_height}")

                # Create object on which to write data
                draw = ImageDraw.Draw(img)
                font = ImageFont.truetype("arial.ttf", size=30)

                # Determine start points of text to be inserted. Determining these x and y coordinates just took 
                # trial-and-error. Ideally, we will test for drawing type (A, B, or C), and have specific values
                # optimized for each type. For now, these work well enough for all three.
                if job_left != None:
                    if job_left < job_top:
                        # Block for A drawing based on Job text
                        job_x = job_left + 220 
                        job_y = job_top - 4 

                        item_x = job_x
                        item_y = job_top + 45 

                        qty_x = job_x
                        qty_y = job_top + 100
                        print("A Type from Job")
                    elif job_left > 1817:
                        # Block for B drawing based on Job text
                        job_x = job_left + 140 
                        job_y = job_top - 10 

                        item_x = job_x
                        item_y = job_top + 20 

                        qty_x = job_x
                        qty_y = job_top + 54 
                        print("B Type from Job")
                    else:
                        # Block for C drawing based on Job text 
                        job_x = job_left + 140
                        job_y = job_top - 10

                        item_x = job_x
                        item_y = job_top + 30

                        qty_x = job_x
                        qty_y = job_top + 70
                        print("C Type from Job")
                elif item_left != None:
                    if item_left < item_top:
                        # Block for A drawing based on Item text
                        job_x = item_left + 220
                        job_y = item_top - 50

                        item_x = job_x 
                        item_y = item_top - 10 

                        qty_x = job_x 
                        qty_y = item_top + 45 
                        print("A Type from Item")
                    elif item_left > 1817:
                        # Block for B drawing based on Item text (not tuned)
                        job_x = item_left + 140 
                        job_y = item_top - 50

                        item_x = job_x 
                        item_y = item_top - 18

                        qty_x = job_x
                        qty_y = item_top + 40
                        print("B Type from Item")
                    else:
                        # Block for C drawing based on Item text (not tuned)
                        job_x = 1
                        job_y = 1

                        item_x = 1
                        item_y = 1

                        qty_x = 1
                        qty_y = 1
                        print("C Type from Item")
                elif qty_left != None:
                    if qty_left < qty_top:
                        # Block for A drawing based on Item text
                        job_x = qty_left + 220
                        job_y = qty_top - 100

                        item_x = job_x
                        item_y = qty_top - 45

                        qty_x = job_x
                        qty_y = qty_top
                        print("A Type from Qty")
                    elif qty_left > 1817:
                        # Block for B drawing based on Item text (not tuned)
                        job_x = 1 
                        job_y = 1

                        item_x = 1
                        item_y = 1

                        qty_x = 1
                        qty_y = 1
                        print("B Type from Qty")
                    else:
                        # Block for C drawing based on Item text (not tuned)
                        job_x = 1
                        job_y = 1

                        item_x = 1
                        item_y = 1

                        qty_x = 1
                        qty_y = 1
                        print("C Type from Qty")
                else:
                    try:
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as f:
                            img.save(f, format="PNG")
                            modified_image_paths.append(f.name)
                    except Exception as error:
                        print(f'Unable to print blank {f.name} to stack.\nError: {error}')
                    print(f"OCR failed for {drawing_filename}.")
                    continue

                # Draw text on the shop copy drawing
                draw.text((job_x, job_y), job_text, font=font, fill=(0, 0, 0))
                draw.text((item_x, item_y), item_text, font=font, fill=(0, 0, 0))
                draw.text((qty_x, qty_y), qty_text, font=font, fill=(0, 0, 0))

                if row[0] in compression_list:
                    #font = ImageFont.truetype("arial.ttf", size=70)
                    #draw.text((300, 300), compression_list[row[0]], font=font, fill=(0, 0, 0))

                    # Create a new image with transparent background to store the rotated text.
                    text_img = Image.new('RGBA', img.size, (255, 255, 255, 0))

                    font = ImageFont.truetype("arial.ttf", size=70)
                    text_draw = ImageDraw.Draw(text_img)

                    # Draw the text onto the text image.
                    if type(compression_list[row[0]]) == str:
                        text_draw.text((300, 300), compression_list[row[0]], font=font, fill=(0, 0, 0))
                    elif type(compression_list[row[0]]) == list:
                        text_draw.text((300, 300), "RUN: " + compression_list[row[0]][0], font=font, fill=(0, 0, 0))
                        text_draw.text((300, 400), "TAP: " + compression_list[row[0]][1], font=font, fill=(0, 0, 0))

                    # Rotate the text image.
                    rotated_text_img = text_img.rotate(30, expand=1)

                    # Create a new image to store the result.
                    result_img = Image.new('RGBA', img.size)

                    # Paste the original image and the rotated text into the result image.
                    result_img.paste(img, (0,0))
                    result_img.paste(rotated_text_img, (0,0), mask=rotated_text_img)

                    # Replace the original image with the result image.
                    img = result_img.convert('RGB')

            except Exception as e:
                print(f"Failed to process {drawing_filename}. Inserting blank drawing. Error: {str(e)}")
                pass

            # The below is a less-than-ideal way of saving but I was running into an issue saving directly as a PDF.
            # Save the image to a temporary file
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as f:
                img.save(f, format="PNG")
                modified_image_paths.append(f.name)

            if self.progress_callback is not None:
                self.progress_callback(progress)
        
        output_directory = "~\\Shop Copies\\"
        output_directory = os.path.expanduser(output_directory)

        if not os.path.exists(output_directory):
            os.makedirs(output_directory)

        output_pdf_path = output_directory + f"CO {self.order_number} SHOP COPY.pdf"

        # Save the modified images as a PDF
        with open(output_pdf_path, "wb") as f:
            f.write(img2pdf.convert(modified_image_paths))

        # Remove temporary image files
        for path in modified_image_paths:
            os.remove(path)

        os.startfile(output_pdf_path)
        
    def set_progress_callback(self, callback):
        self.progress_callback = callback
