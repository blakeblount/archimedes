from datetime import datetime, timedelta
import json
import os
import shutil
import smtplib
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from comtypes.client import CreateObject
from docx import Document
from openpyxl import load_workbook
from jira import JIRA

import config

class PatternWorkOrder:
    def __init__(self, config):
        self.order_number = None
        self.pattern_maker = None
        self.issue_date = datetime.now()
        self.due_date = self.issue_date + timedelta(weeks=12)
        self.orderer = None
        self.part_name = None
        self.drawing_number = None
        self.work_requested = None
        self.pattern_type = None
        self.intended_foundry = None
        self.notes = None

        self.log_book_path = config.get_engineering_log_book_file()
        self.template_path = config.get_pattern_work_order_template_file()
        self.word_folder_path = config.get_pattern_work_order_word_folder()
        self.pdf_folder_path = config.get_pattern_work_order_pdf_folder()
        self.drawings_folder_path = config.get_drawings_folder()

        self.personal_email_address = config.get_personal_email_address()
        self.intertool_email_address = config.get_intertool_email_address()
        self.pro_pattern_email_address = config.get_pro_pattern_email_address()

        self.smtp_server = config.get_smtp_server()
        self.smtp_port = config.get_smtp_port()
        self.personal_email_token = config.get_personal_email_token()

        self.jira_base_url = config.get_jira_base_url()
        self.jira_username = config.get_jira_username()
        self.jira_api_token = config.get_jira_api_token()

    def update_pattern_work_order_fields(self, user_input):
        self.assign_next_order_number()
        self.pattern_maker = user_input["pattern_maker"]
        self.orderer = user_input["orderer"]
        self.part_name = user_input["part_name"]
        self.drawing_number = user_input["drawing_number"]
        self.work_requested = user_input["work_requested"]
        self.pattern_type = user_input["pattern_type"]
        self.intended_foundry = user_input["intended_foundry"]
        self.notes = user_input["notes"]

    def assign_next_order_number(self):
        # Read the Engineering Log Book Excel file
        workbook = load_workbook(self.log_book_path)
        sheet = workbook.active

        # Find the next available order number
        last_order_number_cell = sheet.cell(row=sheet.max_row, column=1)
        next_order_number = int(last_order_number_cell.value.split('-')[1]) + 1
        
        self.order_number = f"P-{next_order_number}"

        #return f"P-{next_order_number}"

    def update_log_book(self):
        workbook = load_workbook(self.log_book_path)
        sheet = workbook.active
        row = sheet.max_row + 1

        sheet.cell(row=row, column=1, value=self.order_number)
        sheet.cell(row=row, column=2, value=self.part_name)
        sheet.cell(row=row, column=3, value=self.drawing_number)
        sheet.cell(row=row, column=4, value=f"{self.work_requested} {self.pattern_type}")
        sheet.cell(row=row, column=5, value=self.pattern_maker)
        sheet.cell(row=row, column=6, value=self.orderer)
        sheet.cell(row=row, column=7, value=self.issue_date.strftime("%m/%d/%Y"))
        sheet.cell(row=row, column=8, value=self.intended_foundry)

        workbook.save(self.log_book_path)

    def generate_word_file(self):
        document = Document(self.template_path)

        # Replace placeholders in the template with pattern_work_order data
        for paragraph in document.paragraphs:
            text = paragraph.text
            text = text.replace("{Order_Number}", self.order_number)
            text = text.replace("{Pattern_Maker}", self.pattern_maker)
            text = text.replace("{Order_Issue_Date}", self.issue_date.strftime("%m/%d/%Y"))
            text = text.replace("{Due_Date}", self.due_date.strftime("%m/%d/%Y"))
            text = text.replace("{Orderer}", self.orderer)
            text = text.replace("{Part_Name}", self.part_name)
            text = text.replace("{Drawing_Number}", self.drawing_number)
            text = text.replace("{Work_Requested}", self.work_requested)
            text = text.replace("{Pattern_Type}", self.pattern_type)
            text = text.replace("{Notes}", self.notes)
            paragraph.text = text

        word_file_name = f"{self.order_number} Pattern Work Order for {self.part_name}.docx"
        word_file_path = os.path.join(self.word_folder_path, word_file_name)
        document.save(word_file_path)

        return word_file_path

    def generate_pdf_file(self, word_file_path):
        word = CreateObject("Word.Application")
        word.Visible = False

        doc = word.Documents.Open(word_file_path)
        pdf_file_name = f"{self.order_number} Pattern Work Order for {self.part_name}.pdf"
        pdf_file_path = os.path.join(self.pdf_folder_path, pdf_file_name)
        doc.SaveAs(pdf_file_path, FileFormat=17)  # 17 represents the PDF file format

        doc.Close()
        word.Quit()

    def send_email(self):
        #Determine emails
        orderer_email = self.personal_email_address
        
        if self.pattern_maker == "Intertool":
           pattern_maker_email = self.intertool_email_address
        elif self.pattern_maker == "Pro Pattern":
           pattern_maker_email = self.pro_pattern_email_address
        
        # Create the email message
        msg = MIMEMultipart()
        msg["From"] = orderer_email
        msg["To"] = pattern_maker_email
        msg["Subject"] = f"{self.order_number} Pattern Work Order for {self.part_name}"

        # Attach the pattern_work_order PDF
        pattern_work_order_pdf_path = os.path.join(self.pdf_folder_path, f"{self.order_number} Pattern Work Order for {self.part_name}.pdf")
        self.attach_file_to_email(msg, pattern_work_order_pdf_path)

        # Attach the drawing PDF
        drawing_pdf_path = os.path.join(self.drawings_folder_path, f"{self.drawing_number}.pdf")
        self.attach_file_to_email(msg, drawing_pdf_path)

        # Add the email body
        body = f"Please see attached Pattern Work Order for {self.part_name}. Let me know if you have any questions."
        msg.attach(MIMEText(body, "plain"))

        # Send the email
        with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
            server.starttls()
            server.login(self.personal_email_address, self.personal_email_token)
            server.sendmail(orderer_email, pattern_maker_email, msg.as_string())

    def attach_file_to_email(self, msg, file_path):
        with open(file_path, "rb") as attachment_file:
            attachment = MIMEBase("application", "octet-stream")
            attachment.set_payload(attachment_file.read())
        encoders.encode_base64(attachment)
        attachment.add_header("Content-Disposition", f"attachment; filename={os.path.basename(file_path)}")
        msg.attach(attachment)

    def attach_pattern_work_order_to_jira_issue(self):
        # Define Jira instance details
        project_key = "PT"

        # Authenticate with Jira
        auth = (self.jira_username, self.jira_api_token)
        jira = JIRA(self.jira_base_url, basic_auth=auth)

        # Define issue summary format
        issue_summary = f"{self.part_name}+"

        # Find issue with the given summary
        issues = jira.search_issues(f'project="{project_key}" AND summary ~ "{issue_summary}"', maxResults=1)

        if issues:
            issue = issues[0]
            pdf_filename = f"{self.order_number} Pattern Work Order for {self.part_name}"

            if os.path.exists(self.pdf_folder_path + "\\" + pdf_filename + ".pdf"):
                # Attach the PDF to the issue
                with open(self.pdf_folder_path + "\\" +  pdf_filename + ".pdf", 'rb') as pdf_file:
                    jira.add_attachment(issue, pdf_file, filename=pdf_filename)
                print(f"Successfully attached {pdf_filename} to Jira issue {issue.key}.")
            else:
                print(f"Could not find PDF file '{pdf_filename}'.")
        else:
            print(f"Could not find Jira issue with summary '{issue_summary}' in project '{project_key}'.")

